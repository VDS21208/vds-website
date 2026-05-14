#!/usr/bin/env python3
"""Extract VDS-Website-Copy-Master.xlsx into JSON files Astro can import."""

import json, os, sys, re
from pathlib import Path
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...", file=sys.stderr)
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

# Resolve paths
script_dir = Path(__file__).parent
project_root = script_dir.parent
data_dir = project_root / "src" / "data"
data_dir.mkdir(parents=True, exist_ok=True)

# Find the XLSX file — try multiple known locations
xlsx_candidates = [
    project_root / "VDS-Website-Copy-Master.xlsx",
    project_root.parent / "VDS-Website-Copy-Master.xlsx",
]
xlsx_path = None
for c in xlsx_candidates:
    if c.exists():
        xlsx_path = c
        break
if not xlsx_path:
    print(f"ERROR: VDS-Website-Copy-Master.xlsx not found. Tried: {xlsx_candidates}", file=sys.stderr)
    sys.exit(1)

print(f"Reading {xlsx_path}")
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

# --- Extract Site Copy tab ---
ws = wb["VDS Site Copy"]
columns = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
pages = []
for r in range(2, ws.max_row + 1):
    row = {}
    for c in range(1, ws.max_column + 1):
        col_name = columns[c - 1]
        if not col_name: continue
        val = ws.cell(row=r, column=c).value
        row[col_name] = val if val is not None else ""
    if row.get("URL"):
        pages.append(row)
print(f"  Site Copy: {len(pages)} pages")

# --- Extract Mega-Menu Nav tab ---
ws = wb["Mega-Menu Nav"]
nav_items = []
for r in range(1, ws.max_row + 1):
    top = ws.cell(row=r, column=1).value
    col = ws.cell(row=r, column=2).value
    order = ws.cell(row=r, column=3).value
    label = ws.cell(row=r, column=4).value
    url = ws.cell(row=r, column=5).value
    typ = ws.cell(row=r, column=6).value
    notes = ws.cell(row=r, column=7).value
    if not top or not isinstance(label, str) or not isinstance(url, str):
        continue
    # Skip section header rows (orange-filled merged cells)
    if str(top).startswith(" ") or "—" in str(top):
        continue
    if top in ("Top-Level Nav", "Section", "Utility (right of nav)", "Announcement bar (top of every page)"):
        continue
    nav_items.append({
        "topLevel": top, "column": col, "order": order,
        "label": label, "url": url, "type": typ, "notes": notes or ""
    })
print(f"  Mega-Menu Nav: {len(nav_items)} items")

# --- Extract Redirects ---
ws = wb["Redirects (Current → New)"]
redirects = []
for r in range(2, ws.max_row + 1):
    current = ws.cell(row=r, column=2).value
    new = ws.cell(row=r, column=3).value
    typ = ws.cell(row=r, column=4).value
    if current and new and "/" in str(current):
        redirects.append({"from": str(current).strip(), "to": str(new).strip(), "type": str(typ or "301")})
print(f"  Redirects: {len(redirects)} rules")

# --- Build helpers ---
def parse_faq(faq_text):
    if not faq_text or not isinstance(faq_text, str):
        return []
    pairs = []
    chunks = faq_text.split("---")
    for chunk in chunks:
        chunk = chunk.strip()
        q_match = re.search(r"^Q:\s*(.+?)(?=\n*A:|\Z)", chunk, re.DOTALL | re.MULTILINE)
        a_match = re.search(r"A:\s*(.+?)$", chunk, re.DOTALL | re.MULTILINE)
        if q_match and a_match:
            pairs.append({"q": q_match.group(1).strip(), "a": a_match.group(1).strip()})
    return pairs

def parse_cta_buttons(text):
    """Parse 'Label1|url1|Label2|url2' into [{label, url}, ...]"""
    if not text or not isinstance(text, str): return []
    parts = text.split("|")
    buttons = []
    for i in range(0, len(parts) - 1, 2):
        label = parts[i].strip()
        url = parts[i+1].strip()
        if label and url:
            buttons.append({"label": label, "url": url})
    return buttons

def parse_hero_cta(text):
    if not text or not isinstance(text, str): return None
    if "|" in text:
        parts = text.split("|", 1)
        return {"label": parts[0].strip(), "url": parts[1].strip()}
    return {"label": text.strip(), "url": "/request-a-quote"}

def parse_stats(text):
    """Parse pipe-separated stats."""
    if not text or not isinstance(text, str): return []
    return [s.strip() for s in text.split("|") if s.strip() and s.strip() != "—"]

# Process each page into a clean structured record
def slugify_part(s):
    return re.sub(r"[^a-z0-9-]", "", s.lower().replace(" ", "-"))

processed_pages = []
for p in pages:
    url = p.get("URL", "").strip()
    if not url: continue
    processed_pages.append({
        "url": url,
        "slug": url.strip("/").split("/")[-1] if url != "/" else "home",
        "parent": p.get("Parent", "").strip(),
        "pageType": p.get("Page Type", "").strip(),
        "status": p.get("Status", "").strip(),
        "seoTitle": p.get("SEO Title (≤60)", "").strip(),
        "metaDescription": p.get("Meta Description (≤155)", "").strip(),
        "announcementBar": p.get("Announcement Bar", "").strip() if p.get("Announcement Bar", "").strip() != "—" else "",
        "h1": p.get("H1", "").strip(),
        "heroEyebrow": p.get("Hero Eyebrow", "").strip(),
        "heroSubhead": p.get("Hero Subhead", "").strip(),
        "heroCTAPrimary": parse_hero_cta(p.get("Hero CTA — Primary", "")),
        "heroCTASecondary": parse_hero_cta(p.get("Hero CTA — Secondary", "")),
        "heroStats": parse_stats(p.get("Hero Stats / Badges", "")),
        "body": p.get("Body (Markdown)", "") if isinstance(p.get("Body (Markdown)"), str) else "",
        "faq": parse_faq(p.get("FAQ (Q&A)", "")),
        "finalCTAHeading": p.get("Final CTA — Heading", "").strip(),
        "finalCTABody": p.get("Final CTA — Body", "").strip(),
        "finalCTAButtons": parse_cta_buttons(p.get("Final CTA — Buttons", "")),
        "schemaType": p.get("Schema Type", "Article").strip(),
        "internalLinks": p.get("Internal Links", "").strip(),
        "notes": p.get("Notes", "").strip() if isinstance(p.get("Notes"), str) else "",
        "canonical": p.get("Canonical URL", "").strip() if isinstance(p.get("Canonical URL"), str) else "",
        "ogTitle": p.get("OG Title", "").strip() if isinstance(p.get("OG Title"), str) else "",
        "ogDescription": p.get("OG Description", "").strip() if isinstance(p.get("OG Description"), str) else "",
        "ogImage": p.get("OG Image", "").strip() if isinstance(p.get("OG Image"), str) else "",
        "ogType": p.get("OG Type", "website").strip() if isinstance(p.get("OG Type"), str) else "website",
        "robots": p.get("Robots Meta", "index, follow").strip() if isinstance(p.get("Robots Meta"), str) else "index, follow",
        "speakable": p.get("Speakable (AEO selectors)", "").strip() if isinstance(p.get("Speakable (AEO selectors)"), str) else "",
        "tags": p.get("Tags / Keywords", "").strip() if isinstance(p.get("Tags / Keywords"), str) else "",
        "author": p.get("Author", "Vision Detection Systems").strip() if isinstance(p.get("Author"), str) else "Vision Detection Systems",
    })

# Group nav items by top level + column
from collections import defaultdict
nav_by_top = defaultdict(lambda: defaultdict(list))
for item in nav_items:
    nav_by_top[item["topLevel"]][item["column"]].append({
        "label": item["label"], "url": item["url"], "type": item["type"], "notes": item["notes"]
    })

# Convert to ordered structure
MENU_DESCRIPTIONS = {
    "/products/sky-guard-mobile-surveillance-trailer": "Flagship trailer · thermal + radar + 4K + LPR",
    "/products/swift-deploy-surveillance-trailer": "Most popular · under-20-min deployment",
    "/products/boundary-guard-pro": "500+ ft fence-line coverage",
    "/products/mobile-surveillance-trailers": "Mobile trailer category overview",
    "/products": "All products and accessories",
    "/products/solar-camera-poles": "Pole-mounted solar-autonomous cameras",
    "/products/livescan-thermal-imaging": "Thermal cameras for night and weather",
    "/products/sentry-ai-monitoring": "AI VisionStream live monitoring software",
    "/products/onboarding-kits": "Pre-flight kits to standardize deployments",
    "/products/portable-camera-units": "Self-contained portable units",
    "/products/accessories": "Mounts, adapters, replacement parts",
    "/platform": "AI VisionStream platform overview",
    "/platform/ai-visionstream": "Cloud-native AI VMS with auto-alerts",
    "/platform/live-monitoring": "24/7 in-house US SOC",
    "/platform/analytics": "Detection, classification, zone analytics",
    "/platform/integrations": "Cameras, NVRs, alarm systems we support",
    "/platform/api-docs": "REST + webhook integration docs",
    "/platform/mobile-app": "iOS/Android live view + alerts",
    "/platform/cybersecurity": "Encryption, MFA, SOC 2 posture",
    "/platform/ndaa-compliance": "NDAA Section 889 details",
    "/industries": "All 17 verticals we protect",
    "/industries/construction-site-security": "Theft + vandalism prevention",
    "/industries/utility-substation-security": "NERC CIP-014 perimeter security",
    "/industries/retail-shopping-center-security": "ORC and after-hours deterrence",
    "/industries/automotive-dealership-security": "Catalytic + inventory protection",
    "/industries/parking-lot-garage-security": "LPR + law-enforcement dispatch",
    "/industries/public-safety-law-enforcement-surveillance": "NDAA + Sourcewell ready",
    "/industries/logistics-warehouse-security": "Yard and 3PL coverage",
    "/industries/cargo-yard-security": "Truck and trailer protection",
    "/industries/event-security": "Short-term rapid deployment",
    "/industries/oil-gas-pipeline-security": "Remote site monitoring",
    "/industries/solar-wind-farm-security": "Copper + asset theft prevention",
    "/industries/critical-infrastructure-security": "NERC + CFATS ready",
    "/industries/government-municipality-security": "Municipal deployments",
    "/industries/cannabis-dispensary-security": "Compliance-grade video retention",
    "/industries/marine-port-security": "Ports + long-range PTZ",
    "/industries/storage-self-storage-security": "LPR and break-in detection",
    "/industries/healthcare-medical-campus-security": "Parking and perimeter",
    "/about-us/why-vision-detection-systems": "What makes VDS different",
    "/about-us": "Our company and mission",
    "/about-us/mission": "Mission and origin story",
    "/about-us/our-plant": "Pikesville, MD US-assembly plant",
    "/about-us/leadership": "Executive team",
    "/about-us/ndaa-compliance": "NDAA Section 889 details",
    "/about-us/cybersecurity": "Encryption, MFA, SOC 2",
    "/about-us/trust-center": "Compliance hub",
    "/about-us/sla-uptime": "Service-level agreement",
    "/about-us/data-privacy": "Privacy and customer rights",
    "/about-us/sustainability": "Solar autonomy + carbon",
    "/about-us/awards": "Awards and certifications",
    "/about-us/careers": "Open roles at VDS",
    "/about-us/press": "News and press releases",
    "/about-us/press-kit": "Logos, photography, brand assets",
    "/about-us/editorial-team": "Our editorial standards",
    "/resources": "Guides, case studies, blog",
    "/resources/blog": "Industry news and field reports",
    "/resources/case-studies": "Real customer outcomes",
    "/resources/faq": "Frequently asked questions",
    "/resources/news": "Press and announcements",
    "/resources/videos": "Product demos and walkthroughs",
    "/resources/webinars": "Live and on-demand webinars",
    "/resources/state-of-mobile-surveillance-2026": "Our 2026 industry report",
    "/resources/camera-sensor-types": "Camera sensor types guide",
    "/guides": "Buyer's playbook and how-to guides",
    "/guides/pricing-guide-2026": "2026 pricing guide",
    "/guides/roi-calculator": "Estimate your VDS ROI",
    "/glossary": "Industry terms, sensors, hardware",
    "/answers": "Quick answers to common questions",
    "/partners": "Dealer & integrator program",
    "/partners/dealer-program": "Become a VDS dealer",
    "/partners/integrator-program": "Integrator partnership details",
    "/specifiers": "Spec sheets and architect resources",
    "/compare": "Honest competitor comparisons",
    "/contact-us": "Talk to our team",
    "/pricing": "Transparent pricing",
    "/products/guard-booths": "Manned guard booth solutions",
    "/products/medical-cooling-stations": "Mobile medical and cooling units",
    "/products/light-trailers": "Mobile lighting trailers",
    "/products/license-plate-recognition": "LPR-equipped trailers",
    "/platform/ai-video-analytics": "Object, person, vehicle detection",
    "/platform/cloud-video-surveillance": "Cloud video storage + playback",
    "/platform/alerts-reporting": "Real-time alerts and incident reports",
    "/industries/oil-gas-security": "Remote pipeline + facility coverage",
    "/industries/hospitality-healthcare-security": "Hotels, hospitals, campuses",
    "/industries/school-campus-security": "K-12 and university perimeters",
    "/industries/commercial-property-security": "Multi-tenant + property mgmt",
    "/industries/warehouse-logistics-security": "Yard, 3PL, distribution centers",
    "/guides/construction-theft-prevention": "Stopping copper + equipment theft",
    "/guides/ndaa-compliance": "NDAA Section 889 buying guide",
    "/guides/utility-substation-security": "Substation perimeter playbook",
    "/guides/retail-orc": "Organized retail crime defense",
    "/partners/become-a-dealer": "Join the VDS dealer network",
    "/partners/white-label": "White-label program",
    "/partners/resources": "Sales tools and collateral",
    "/partners/financing": "Customer financing options",
    "/partners/find-a-dealer": "Locate a local dealer",
    "/integrations": "Integrations directory",
    "/integrations/genetec": "Genetec Security Center integration",
    "/integrations/milestone": "Milestone XProtect integration",
    "/platform/soc": "24/7 in-house US Security Operations Center",
    "/platform/lpr": "LPR · 95%+ accuracy, <3s alerts",
    "/platform/api-sdk-access": "Developer + integrator API/SDK",
    "/platform/cloud-security": "NDAA, TLS 1.3, AES-256",
    "/platform/pricing": "Platform subscription tiers",
    "/industries/insurance-carriers": "Insurance channel buyers",
}

def _fill_notes(items):
    for it in items:
        if not it.get("notes") or len(str(it.get("notes", "")).strip()) < 3:
            url = (it.get("url") or "").rstrip("/")
            desc = MENU_DESCRIPTIONS.get(url) or MENU_DESCRIPTIONS.get(url + "/")
            if desc:
                it["notes"] = desc
    return items

mega_menu = []
TOP_ORDER = ["Products", "Platform", "Industries", "Why VDS", "Resources", "Partners"]
for top in TOP_ORDER:
    if top in nav_by_top:
        cols = nav_by_top[top]
        mega_menu.append({
            "label": top,
            "url": "/" + slugify_part(top) if top != "Why VDS" else "/compare",
            "primary": _fill_notes(cols.get("Primary", cols.get("Tools", cols.get("Dealer Program", [])))),
            "supporting": _fill_notes(cols.get("Supporting", cols.get("Library", cols.get("Channel & Integrations", [])))),
            "featured": cols.get("Featured", []),
            "cta": cols.get("CTA", []),
        })

# Write JSON files
with open(data_dir / "pages.json", "w") as f:
    json.dump(processed_pages, f, indent=2)
with open(data_dir / "nav.json", "w") as f:
    json.dump(mega_menu, f, indent=2)
with open(data_dir / "redirects.json", "w") as f:
    json.dump(redirects, f, indent=2)

# Build URL → page index for fast lookup in Astro
url_index = {p["url"]: p for p in processed_pages}
with open(data_dir / "page-index.json", "w") as f:
    json.dump(url_index, f, indent=2)

# Build collections by page type for dynamic routes
collections = defaultdict(list)
for p in processed_pages:
    pt = (p["pageType"] or "").lower()
    url = p["url"]
    # Map URLs to collection types
    if url.startswith("/products/") and url != "/products":
        collections["products"].append(p)
    elif url.startswith("/industries/") and url != "/industries":
        collections["industries"].append(p)
    elif url.startswith("/compare/") and url != "/compare":
        collections["compare"].append(p)
    elif url.startswith("/platform/") and url != "/platform":
        collections["platform"].append(p)
    elif url.startswith("/services/") and url != "/services":
        collections["services"].append(p)
    elif url.startswith("/locations/mobile-surveillance-trailers/"):
        collections["city-locations"].append(p)
    elif url.startswith("/locations/") and url != "/locations":
        collections["state-locations"].append(p)
    elif url.startswith("/solutions/") and url != "/solutions":
        collections["solutions"].append(p)
    elif url.startswith("/guides/") and url != "/guides":
        collections["guides"].append(p)
    elif url.startswith("/use-cases/") and url != "/use-cases":
        collections["use-cases"].append(p)
    elif url.startswith("/resources/blog/"):
        collections["blog"].append(p)
    elif url.startswith("/resources/case-studies/"):
        collections["case-studies"].append(p)
    elif url.startswith("/integrations/") and url != "/integrations":
        collections["integrations"].append(p)
    elif url.startswith("/about-us/") and url != "/about-us":
        collections["about"].append(p)
    elif url.startswith("/partners/") and url != "/partners":
        collections["partners"].append(p)
    elif url.startswith("/pricing/") and url != "/pricing":
        collections["pricing"].append(p)

for col_name, items in collections.items():
    print(f"  Collection '{col_name}': {len(items)} pages")
    with open(data_dir / f"collection-{col_name}.json", "w") as f:
        json.dump(items, f, indent=2)

# Save a manifest of all collections
manifest = {name: len(items) for name, items in collections.items()}
manifest["total"] = len(processed_pages)
with open(data_dir / "manifest.json", "w") as f:
    json.dump(manifest, f, indent=2)

print(f"\nWrote data to {data_dir}")
print(f"Total pages indexed: {len(processed_pages)}")
